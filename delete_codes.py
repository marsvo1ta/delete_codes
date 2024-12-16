from openpyxl import load_workbook
from api_request import ApiRequest


api = ApiRequest()


def process_excel_column_a(file):
    workbook = load_workbook(file)
    sheet = workbook.active

    result = []
    ids = []
    for row in sheet.iter_rows(min_col=1, max_col=1, values_only=True):
        cell_value: str = row[0]
        result.append(cell_value)
        ids_from_code = api.get_ids_from_v2(cell_value)
        ids.append(ids_from_code)
        api.bulk_delete(ids_from_code)
    workbook.close()
    return result, ids



